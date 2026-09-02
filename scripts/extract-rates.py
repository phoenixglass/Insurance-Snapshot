#!/usr/bin/env python3
"""Regenerate src/data/rates.js from a Deposit Calculator workbook.

The rate table in the app is a mirror of the workbook's `Vlookup` sheet, and
the header of the generated file has always said to regenerate rather than
hand-edit it. This is the thing that does that.

    pip install openpyxl
    python3 scripts/extract-rates.py "Deposit Calculator 2026 Revised.xlsx"

What it reads, all of it addressed the way the workbook's own named ranges are:

    CarrierList      Vlookup!B4:B61    carrier names, with A holding the network
    CPTCodeList      Vlookup!C3:DT3    codes, with row 2 holding the description
    TreatmentSeqs    Vlookup!EK3:EK45  the pathway dropdown

A blank rate cell is a real state — the workbook marks it amber and tells the
user to estimate from a similar plan — so a blank is left out of the carrier's
row rather than written as zero. A literal 0 is dropped the same way: the sheet
uses it as a placeholder, and a $0 allowed amount would silently zero a line
the app should be flagging as unpriced.

Anything the app keeps beyond the workbook (treatment sequences it still
offers, rates the workbook has wrong) is layered on elsewhere and is not this
script's business — see EXTRA_TREATMENT_SEQUENCES below and
src/data/rateCorrections.js.
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# Pathways the app offers that the 2026 revision of the workbook dropped.
# Kept because a step-down you cannot select is a worse failure than an extra
# option in a list, and the estimator prices any pathway built from these
# levels. Written here rather than in the generated file so a re-export does
# not silently take them away again.
EXTRA_TREATMENT_SEQUENCES = {
    'Residential > PHP': 'Residential',
    'Residential > OP': 'Residential > IOP',
    'OPWM > Residential > IOP': 'OPWM > Residential > PHP',
    'OPWM > Residential > IOP > OP': 'OPWM > Residential > PHP > IOP',
}

HEADER = '''// ─────────────────────────────────────────────────────────────────────────────
// Reference data extracted from the {book} workbook's
// `Vlookup` sheet. This is the same table the workbook's INDEX/MATCH formulas
// read: one contracted/allowed rate per carrier per CPT-HCPCS code.
//
// A missing rate is a real state, not a zero. The workbook marks those cells
// amber and instructs the user to estimate from a similar plan, so a code the
// carrier has no rate for is absent from its row and the app says so rather
// than pricing the service at $0. A cell holding a literal 0 is dropped the
// same way — the sheet uses it as a placeholder, and a $0 allowed amount would
// quietly zero a line instead of flagging it.
//
// Generated file — run `python3 scripts/extract-rates.py <workbook.xlsx>`
// rather than editing by hand. Rates the workbook has wrong are corrected in
// `rateCorrections.js`, which is laid over this table at lookup time.
// ─────────────────────────────────────────────────────────────────────────────
'''


def money(v):
    """A rate to the cent.

    Six cells in the sheet carry sub-cent precision left over from an average
    — 173.555, 164.225 — and an allowed amount is a dollar figure, not the tail
    of a division. Rounding here rather than at every point of use keeps one
    number in circulation instead of two that differ in the third decimal.
    """
    return int(v) if float(v) == int(v) else round(float(v), 2)


def js(s):
    return '"' + str(s).replace('\\', '\\\\').replace('"', '\\"') + '"'


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Vlookup']

    # The named ranges the workbook itself uses, so a resized sheet is followed
    # rather than guessed at.
    names = {n: d.value for n, d in wb.defined_names.items()}

    def bounds(defined, fallback):
        ref = names.get(defined, fallback)
        body = ref.split('!')[-1].replace('$', '')
        a, b = body.split(':')
        return a, b

    from openpyxl.utils import range_boundaries, column_index_from_string
    c0, r0, c1, r1 = range_boundaries(':'.join(bounds('CarrierList', 'B4:B61')))
    carrier_rows = range(r0, r1 + 1)
    k0, kr, k1, _ = range_boundaries(':'.join(bounds('CPTCodeList', 'C3:DT3')))
    code_cols = range(k0, k1 + 1)
    s0, sr0, s1, sr1 = range_boundaries(':'.join(bounds('TreatmentSequences', 'EK3:EK45')))

    codes = [(c, str(ws.cell(kr, c).value).strip(), str(ws.cell(kr - 1, c).value or '').strip())
             for c in code_cols]
    carriers = [(r, str(ws.cell(r, c0).value).strip(), str(ws.cell(r, c0 - 1).value or '').strip())
                for r in carrier_rows if ws.cell(r, c0).value]

    sequences = [ws.cell(r, s0).value for r in range(sr0, sr1 + 1)]
    sequences = [str(s).strip() for s in sequences if s]
    for extra, after in EXTRA_TREATMENT_SEQUENCES.items():
        if extra not in sequences:
            sequences.insert(sequences.index(after) + 1 if after in sequences else len(sequences), extra)

    # Uploads arrive with a hash prefix and underscores; the header should name
    # the workbook the way a person would.
    book = re.sub(r'^[0-9a-f]{6,}-', '', Path(path).stem).replace('_', ' ').strip()
    out = [HEADER.format(book=re.sub(r'\s+', ' ', book))]

    out.append('export const CARRIERS = [')
    for _, name, network in carriers:
        out.append(f'  {{ name: {js(name)}, network: {js(network)} }},')
    out.append(']\n')

    out.append('// Every CPT / HCPCS code the workbook prices, in workbook order.')
    out.append('export const CODES = [')
    for _, code, desc in codes:
        out.append(f'  {{ code: {js(code)}, description: {js(desc)} }},')
    out.append(']\n')

    out.append('// carrier name → { code: allowed rate }. Codes the carrier has no rate for are')
    out.append('// absent from the object.')
    out.append('export const RATES = {')
    for r, name, _ in carriers:
        cells = []
        for c, code, _d in codes:
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and float(v) != 0:
                cells.append(f'{js(code)}: {money(v)}')
        out.append(f'  {js(name)}: {{ ' + ', '.join(cells) + ' },' if cells else f'  {js(name)}: {{}},')
    out.append('}\n')

    # The workbook's own Grand Total row, one row under the carrier list. It is
    # a pivot average over the source data rather than a mean of the visible
    # cells, so it is read rather than recomputed — a mean of this sheet's own
    # rows would be a different number wearing the same label.
    out.append("// The workbook's cross-carrier average for each primary code, read from its")
    out.append('// own Grand Total row. Offered as a visible benchmark when a carrier has no')
    out.append('// rate on file — never substituted silently into a quote.')
    out.append('export const BENCHMARK_RATES = {')
    total_row = r1 + 1
    if str(ws.cell(total_row, c0).value or '').strip().lower() != 'grand total':
        sys.exit(f'expected a Grand Total row at row {total_row}; found '
                 f'{ws.cell(total_row, c0).value!r}. Check the sheet before regenerating.')
    for c, code, _d in codes:
        v = ws.cell(total_row, c).value
        if isinstance(v, (int, float)) and float(v) != 0:
            out.append(f'  {js(code)}: {round(float(v), 2)},')
    out.append('}\n')

    out.append(f'// The {len(sequences)} treatment-sequence orders the app offers — the workbook\'s')
    out.append(f'// dropdown plus {len(EXTRA_TREATMENT_SEQUENCES)} pathways kept past the 2026 revision that dropped them')
    out.append('// (see EXTRA_TREATMENT_SEQUENCES in scripts/extract-rates.py). A level of care')
    out.append('// is active for an estimate only when it appears in the selected sequence.')
    out.append('export const TREATMENT_SEQUENCES = [')
    for s in sequences:
        out.append(f'  {js(s)},')
    out.append(']\n')

    out.append('// Searchable service labels for the rate lookup, paired to their code.')
    out.append('export const CODE_SEARCH_LABELS = [')
    for _c, code, desc in codes:
        out.append(f'  {{ code: {js(code)}, label: {js(desc)} }},')
    out.append(']')

    print(f'carriers {len(carriers)}  codes {len(codes)}  sequences {len(sequences)}', file=sys.stderr)
    return '\n'.join(out) + '\n'


if __name__ == '__main__':
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    target = Path(__file__).resolve().parent.parent / 'src' / 'data' / 'rates.js'
    target.write_text(main(sys.argv[1]))
    print(f'wrote {target}', file=sys.stderr)
